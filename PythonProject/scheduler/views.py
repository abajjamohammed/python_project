# ============= IMPORTS =============
# Django Core
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Q
from django.utils import timezone

# Python Standard Library
import csv
import json
from datetime import datetime, timedelta

# Excel/PDF Libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io

# Local Imports
from .models import (
    Room, Course, ReservationRequest, User, ScheduledSession, 
    Filiere, TeacherUnavailability
)
from .forms import (
    ReservationForm, CourseForm, TeacherForm, TeacherEditForm,
    RoomSearchForm, SessionForm, TeacherUnavailabilityForm, ProfileForm
)
from .utils import TimetableAlgorithm


# ============= VIEWS =============

@login_required
def dashboard_router(request):
    """The 'Home' page that redirects users based on their role"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.user.role == 'A':
        return redirect('admin_dashboard')
    elif request.user.role == 'T':
        return redirect('teacher_dashboard')
    else:
        return redirect('student_dashboard')


def admin_dashboard(request):
    # 1. Fetch Real Data from Database
    total_students = User.objects.filter(role='S').count()
    total_teachers = User.objects.filter(role='T').count()
    total_rooms = Room.objects.count()
    
    # Get the actual reservation objects (not just the count)
    pending_reservations = ReservationRequest.objects.filter(status='PENDING')
    pending_requests = pending_reservations.count()

    # 2. Data for the Chart
    sessions_per_day = ScheduledSession.objects.values('day').annotate(count=Count('id'))
    
    if not sessions_per_day:
        chart_labels = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        chart_data = [0, 0, 0, 0, 0]
    else:
        chart_labels = [item['day'] for item in sessions_per_day]
        chart_data = [item['count'] for item in sessions_per_day]

    # 3. Pack everything into context
    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_rooms': total_rooms,
        'pending_requests': pending_requests,
        'pending_reservations': pending_reservations,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    }

    return render(request, 'scheduler/dashboard.html', context)


@login_required
def teacher_timetable(request):
    """Dedicated timetable page for teachers with Rowspan logic"""
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    hours = range(8, 19)
    
    # 1. Get official classes (ScheduledSession)
    sessions = ScheduledSession.objects.filter(
        course__teacher=request.user
    ).select_related('course', 'room', 'course__filiere')

    # --- Fetch Approved Reservations ---
    reservations = ReservationRequest.objects.filter(
        teacher=request.user,
        status='APPROVED'
    )
    
    # 2. Calculate stats
    total_sessions = sessions.count() + reservations.count() # Update count
    
    # Calculate weekly hours (Sessions + Reservations)
    session_hours = sum([(s.end_hour - s.start_hour) for s in sessions])
    reservation_hours = sum([(r.end_hour - r.start_hour) for r in reservations])
    weekly_hours = session_hours + reservation_hours
    
    now = timezone.localtime()
    today_name = now.strftime('%A')
    
    today_sessions_count = sessions.filter(day=today_name).count() + reservations.filter(day=today_name).count()
    
    # Next class logic (Check both sessions and reservations)
    next_class = sessions.filter(day=today_name, start_hour__gte=now.hour).order_by('start_hour').first()
    if not next_class:
        # If no official class, check if there is a reservation
        next_class = reservations.filter(day=today_name, start_hour__gte=now.hour).order_by('start_hour').first()
        # If found, we need to hack it slightly so the template can read 'course.name'
        if next_class:
            class TempCourse:
                name = next_class.reason
                get_session_type_display = "Reservation"
            next_class.course = TempCourse()

    # 3. Build a quick lookup dictionary
    session_map = {}
    
    # Add Official Sessions
    for s in sessions:
        session_map[(s.day, s.start_hour)] = s
        
    for r in reservations:
        # We need to make the Reservation look like a Session for the HTML template
        # The template expects: session.course.name, session.course.get_session_type_display, etc.
        class PlaceholderCourse:
            name = r.reason  # Use the 'Reason' as the Course Name
            session_type = "RES" 
            get_session_type_display = "Reservation" # Label for the badge
            group = None     # No group
            filiere = None   # No filiere
            
            # Helper to prevent template errors
            class FilierePlaceHolder:
                code = "PERSO"
            filiere = FilierePlaceHolder()

        r.course = PlaceholderCourse() # Attach this fake course to the reservation
        
        # Add to the map (It will appear in the grid)
        session_map[(r.day, r.start_hour)] = r

    # 4. Build the Grid Matrix
    timetable_data = []
    skip_slots = set()

    for h in hours:
        row = {'hour': h, 'slots': []}
        for d in days:
            if (d, h) in skip_slots:
                row['slots'].append({'type': 'skipped'})
                continue

            session = session_map.get((d, h))
            
            if session:
                duration = session.end_hour - session.start_hour
                for i in range(1, duration):
                    skip_slots.add((d, h + i))
                
                row['slots'].append({
                    'type': 'session',
                    'session': session,
                    'rowspan': duration
                })
            else:
                row['slots'].append({'type': 'empty'})
        
        timetable_data.append(row)

    context = {
        'timetable_data': timetable_data,
        'days': days,
        'next_class': next_class,
        'today_sessions_count': today_sessions_count,
        'sessions_count': total_sessions,
        'weekly_hours': weekly_hours,
        'user_name': request.user.get_full_name() or request.user.username,
    }
    return render(request, 'scheduler/teacher_timetable.html', context)


@login_required
def student_timetable(request):
    """Student timetable with proper rowspan handling"""
    student_group = request.user.student_group

    if not student_group:
        return render(request, 'scheduler/student_timetable.html', {
            'error': 'You are not assigned to any group.'
        })

    # Get sessions
    sessions = ScheduledSession.objects.filter(
        Q(course__group=student_group) | 
        Q(course__filiere=student_group.filiere, course__group__isnull=True)
    ).select_related('course', 'room', 'course__teacher', 'course__filiere').order_by('day', 'start_hour')

    # Prepare data structure
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    hours = range(8, 19)
    
    session_map = {}
    for s in sessions:
        session_map[(s.day, s.start_hour)] = s
    
    skip_slots = set()
    timetable_data = []
    
    for h in hours:
        row = {'hour': h, 'slots': []}
        for d in days:
            if (d, h) in skip_slots:
                row['slots'].append({'type': 'skipped'})
                continue

            session = session_map.get((d, h))
            
            if session:
                duration = session.end_hour - session.start_hour
                for i in range(1, duration):
                    skip_slots.add((d, h + i))
                
                row['slots'].append({
                    'type': 'session',
                    'session': session,
                    'rowspan': duration
                })
            else:
                row['slots'].append({'type': 'empty'})
        
        timetable_data.append(row)
    
    # Calculate stats
    total_sessions = sessions.count()
    weekly_hours = sum([(s.end_hour - s.start_hour) for s in sessions])
    
    # Find next class
    today_name = datetime.now().strftime('%A')
    current_hour = datetime.now().hour
    
    next_class = sessions.filter(day=today_name, start_hour__gte=current_hour).order_by('start_hour').first()
    if not next_class:
        next_day_index = (days.index(today_name) + 1) % len(days) if today_name in days else 0
        next_class = sessions.filter(day=days[next_day_index]).order_by('start_hour').first()
    
    today_sessions_count = sessions.filter(day=today_name).count()

    context = {
        'timetable_data': timetable_data,
        'days': days,
        'hours': hours,
        'next_class': next_class,
        'today_sessions_count': today_sessions_count,
        'total_sessions': total_sessions,
        'weekly_hours': weekly_hours,
        'student_group': student_group,
        'student_name': request.user.get_full_name() or request.user.username,
    }
    
    return render(request, 'scheduler/student_timetable.html', context)


@login_required
def generate_timetable(request):
    # Security: Ensure only Admins can do this
    if not request.user.is_authenticated or request.user.role != 'A': 
        messages.error(request, "Accès refusé.")
        return redirect('login')

    # 1. Clear existing schedule
    ScheduledSession.objects.all().delete()

    # 2. Trigger the Algorithm
    algo = TimetableAlgorithm()
    unscheduled = algo.generate_timetable()

    # 3. Success/Warning message
    if not unscheduled:
        messages.success(request, "L'emploi du temps a été généré avec succès !")
    else:
        messages.warning(request, f"Généré, mais impossible de placer : {', '.join(unscheduled)}")

    return redirect('admin_dashboard')


@login_required
def make_reservation(request):
    """Permet à un prof de faire une demande"""
    if request.method == 'POST':
        form = ReservationForm(request.POST)
        if form.is_valid():
            reservation = form.save(commit=False)
            reservation.teacher = request.user
            reservation.save()
            messages.success(request, "Demande envoyée avec succès !")
            return redirect('dashboard')
    else:
        form = ReservationForm()
    
    return render(request, 'scheduler/make_reservation.html', {
        'form': form,
        'errors': form.errors 
    })


@login_required
def approve_reservations(request):
    """Liste les demandes en attente pour l'admin"""
    requests = ReservationRequest.objects.filter(status='PENDING')
    return render(request, 'scheduler/approve_reservations.html', {'requests': requests})


@login_required
def process_request(request, req_id, action):
    """Traite l'action Accepter ou Refuser"""
    reservation = get_object_or_404(ReservationRequest, id=req_id)
    
    if action == 'approve':
        reservation.status = 'APPROVED'
        reservation.save()
        messages.success(request, "Réservation approuvée.")
    elif action == 'reject':
        reservation.status = 'REJECTED'
        reservation.save()
        messages.error(request, "Réservation rejetée.")
    
    return redirect('approve_reservations')


def teacher_list(request):
    teachers = User.objects.filter(role='T')
    return render(request, 'scheduler/teacher_list.html', {'teachers': teachers})


def add_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('course_list')
    else:
        form = CourseForm()
    
    return render(request, 'scheduler/add_course.html', {'form': form})


def timetable_view(request):
    """Admin timetable page - HTML table version"""
    # 1. Get all filières for dropdown
    filieres = Filiere.objects.all().order_by('level', 'code')
    
    # 2. Get selected filière or show all
    selected_filiere_id = request.GET.get('filiere')
    selected_filiere = None
    
    if selected_filiere_id:
        try:
            selected_filiere = Filiere.objects.get(id=selected_filiere_id)
            sessions = ScheduledSession.objects.filter(
                course__filiere=selected_filiere
            ).select_related('course', 'room', 'course__teacher', 'course__filiere', 'course__group')
        except Filiere.DoesNotExist:
            sessions = ScheduledSession.objects.all().select_related(
                'course', 'room', 'course__teacher', 'course__filiere', 'course__group'
            )
    else:
        sessions = ScheduledSession.objects.all().select_related(
            'course', 'room', 'course__teacher', 'course__filiere', 'course__group'
        )
    
    # 3. Days and hours
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    hours = range(8, 19)
    
    # 4. Create session map
    session_map = {}
    skip_slots = set()
    
    for session in sessions:
        session_map[(session.day, session.start_hour)] = session
    
    # 5. Build timetable grid
    timetable_data = []
    for h in hours:
        row = {'hour': h, 'slots': []}
        for d in days:
            if (d, h) in skip_slots:
                row['slots'].append({'type': 'skipped'})
                continue
            
            session = session_map.get((d, h))
            
            if session:
                duration = session.end_hour - session.start_hour
                for i in range(1, duration):
                    skip_slots.add((d, h + i))
                
                row['slots'].append({
                    'type': 'session',
                    'session': session,
                    'rowspan': duration
                })
            else:
                row['slots'].append({'type': 'empty'})
        
        timetable_data.append(row)
    
    # 6. Calculate stats
    total_sessions = sessions.count()
    weekly_hours = sum([(s.end_hour - s.start_hour) for s in sessions])
    today_name = datetime.now().strftime('%A')
    today_sessions_count = sessions.filter(day=today_name).count()
    
    context = {
        'filieres': filieres,
        'selected_filiere': selected_filiere,
        'timetable_data': timetable_data,
        'days': days,
        'hours': hours,
        'total_sessions': total_sessions,
        'weekly_hours': weekly_hours,
        'today_sessions_count': today_sessions_count,
    }
    
    return render(request, 'scheduler/timetable.html', context)


def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('teacher_list')
    else:
        form = TeacherForm()

    return render(request, 'scheduler/add_teacher.html', {'form': form})


@login_required
def teacher_dashboard(request):
    # Get teacher's sessions
    sessions = ScheduledSession.objects.filter(course__teacher=request.user).select_related('course', 'room')
    
    # Get today's sessions (In English, to match the database)
    now = timezone.localtime()
    today_name = now.strftime('%A') # e.g., "Wednesday"
    
    todays_sessions = sessions.filter(day=today_name).order_by('start_hour')
    # --------------------------
    
    # Get reservation requests
    my_reqs = ReservationRequest.objects.filter(teacher=request.user).order_by('-id')
    
    # Calculate stats
    total_courses = sessions.values('course').distinct().count()
    weekly_hours = sum([(s.end_hour - s.start_hour) for s in sessions])
    pending_requests_count = my_reqs.filter(status='PENDING').count()
    
    context = {
        'sessions': sessions,
        'todays_sessions': todays_sessions,
        'todays_sessions_count': todays_sessions.count(),
        'my_reqs': my_reqs,
        'total_courses': total_courses,
        'weekly_hours': weekly_hours,
        'pending_requests_count': pending_requests_count,
        'days': ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
        'hours': [(8, 10), (10, 12), (14, 16), (16, 18)]
    }
    return render(request, 'scheduler/teacher_dashboard.html', context)


@login_required
def student_dashboard(request):
    # Setup grid variables
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    hours = range(8, 19)
    
    user = request.user
    student_group = user.student_group
    
    # Safety check
    if not student_group:
        return render(request, 'scheduler/student_dashboard.html', {
            'days': days, 'hours': hours, 'sessions': [],
            'error_message': 'No group assigned.'
        })

    # Get sessions (Both TD/TP for the group and CM for the whole filiere)
    sessions = ScheduledSession.objects.filter(
        Q(course__group=student_group) | 
        Q(course__filiere=student_group.filiere, course__session_type='CM')
    ).select_related('course', 'room', 'course__teacher')
    for session in sessions:
    # Calculate how many hours the session lasts
        session.duration = session.end_hour - session.start_hour
    # --- ENHANCED NEXT CLASS LOGIC ---
    now = timezone.localtime()
    today_name = now.strftime('%A')
    current_hour = now.hour
    
    # We create an ordered list starting from Today
    day_sequence = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    # If it's Sunday, we want to start looking from Monday
    if today_name in day_sequence:
        start_index = day_sequence.index(today_name)
        # Reorder list so it starts with today (e.g., [Mon, Tue, Wed...] or [Wed, Thu, Fri... Mon, Tue])
        ordered_days = day_sequence[start_index:] + day_sequence[:start_index]
    else:
        ordered_days = day_sequence # It's Sunday, look from Monday onwards

    next_class = None
    status_label = "NEXT UP"

    for day in ordered_days:
        if day == today_name:
            # If checking today, only look for classes that haven't ended yet
            found = sessions.filter(day=day, end_hour__gt=current_hour).order_by('start_hour').first()
        else:
            # If checking a future day, just take the first class of that day
            found = sessions.filter(day=day).order_by('start_hour').first()
        
        if found:
            next_class = found
            # Update label if the class is not today
            if day != today_name:
                status_label = f"NEXT CLASS: {day.upper()}"
            break

    context = {
        'student_group': student_group,
        'sessions': sessions,
        'next_class': next_class,
        'status_label': status_label,
        'days': days,
        'hours': hours,
        'today_name': today_name,
        'course_count': sessions.values('course').distinct().count(),
    }
    return render(request, 'scheduler/student_dashboard.html', context)

@login_required
def export_timetable_csv(request):
    """Generates a CSV file of the timetable for Excel export"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="university_timetable.xl"'

    writer = csv.writer(response)
    writer.writerow(['Course Name', 'Teacher', 'Room', 'Day', 'Start Time', 'End Time'])

    # Determine data based on user role
    if request.user.role == 'A':
        sessions = ScheduledSession.objects.all().select_related('course', 'room', 'course__teacher')
    elif request.user.role == 'T':
        sessions = ScheduledSession.objects.filter(course__teacher=request.user).select_related('course', 'room')
    else:
        sessions = ScheduledSession.objects.filter(course__group_name=request.user.student_group).select_related('course', 'room')

    # Fill CSV
    for session in sessions:
        writer.writerow([
            session.course.name,
            session.course.teacher.username,
            session.room.name,
            session.day,
            f"{session.start_hour}:00",
            f"{session.end_hour}:00"
        ])

    return response


@login_required
def my_reservations(request):
    """Teacher views their own reservation history"""
    my_reqs = ReservationRequest.objects.filter(teacher=request.user).order_by('-id')
    return render(request, 'scheduler/my_reservations.html', {'my_reqs': my_reqs})


def custom_logout(request):
    logout(request)
    return redirect('login')


@login_required
def student_classes(request):
    """List of all courses for the student"""
    student_group = request.user.student_group
    
    if not student_group:
        return render(request, 'scheduler/student_classes.html', {
            'courses': [],
            'error': 'No group assigned.'
        })

    courses = Course.objects.filter(
        Q(group=student_group) | 
        Q(filiere=student_group.filiere, group__isnull=True)
    ).select_related('teacher', 'filiere').order_by('name')
    
    context = {
        'courses': courses,
        'student_group': student_group,
    }
    return render(request, 'scheduler/student_classes.html', context)


def add_session(request):
    if request.method == 'POST':
        form = SessionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('timetable') 
    else:
        form = SessionForm()
    
    return render(request, 'scheduler/add_session.html', {'form': form})


def edit_teacher(request, teacher_id):
    teacher = get_object_or_404(User, pk=teacher_id)
    
    if request.method == 'POST':
        form = TeacherEditForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('teacher_list')
    else:
        form = TeacherEditForm(instance=teacher)
    
    return render(request, 'scheduler/add_teacher.html', {'form': form, 'title': 'Edit Teacher'})


def delete_teacher(request, teacher_id):
    teacher = get_object_or_404(User, pk=teacher_id)
    
    if request.method == 'POST':
        teacher.delete()
        return redirect('teacher_list')
    
    return render(request, 'scheduler/confirm_delete.html', {'teacher': teacher})


@login_required
def find_rooms(request):
    """Search for available rooms"""
    available_rooms = []
    search_performed = False
    
    if request.method == 'GET' and any(request.GET.values()):
        search_performed = True
        day = request.GET.get('day')
        start_hour = request.GET.get('start_hour')
        end_hour = request.GET.get('end_hour')
        min_capacity = request.GET.get('min_capacity')
        
        rooms = Room.objects.all()
        
        if min_capacity:
            rooms = rooms.filter(capacity__gte=int(min_capacity))
        
        if day and start_hour and end_hour:
            occupied_by_sessions = ScheduledSession.objects.filter(
                day=day,
                start_hour__lt=int(end_hour),
                end_hour__gt=int(start_hour)
            ).values_list('room_id', flat=True)
            
            occupied_by_reservations = ReservationRequest.objects.filter(
                day=day,
                status='APPROVED',
                start_hour__lt=int(end_hour),
                end_hour__gt=int(start_hour)
            ).values_list('room_id', flat=True)
            
            occupied_room_ids = set(occupied_by_sessions) | set(occupied_by_reservations)
            available_rooms = rooms.exclude(id__in=occupied_room_ids)
        else:
            available_rooms = rooms
    
    context = {
        'available_rooms': available_rooms,
        'search_performed': search_performed,
        'days': ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
        'hours': range(8, 19),
    }
    return render(request, 'scheduler/find_rooms.html', context)


@login_required
def manage_unavailability(request):
    """Teacher can view and manage their unavailability"""
    unavailabilities = TeacherUnavailability.objects.filter(teacher=request.user).order_by('day', 'start_hour')
    
    if request.method == 'POST':
        form = TeacherUnavailabilityForm(request.POST)
        if form.is_valid():
            unavailability = form.save(commit=False)
            unavailability.teacher = request.user
            unavailability.save()
            messages.success(request, "Unavailability added successfully!")
            return redirect('manage_unavailability')
    else:
        form = TeacherUnavailabilityForm()
    
    context = {
        'form': form,
        'unavailabilities': unavailabilities,
    }
    return render(request, 'scheduler/manage_unavailability.html', context)


@login_required
def delete_unavailability(request, unavail_id):
    """Delete an unavailability entry"""
    unavailability = get_object_or_404(TeacherUnavailability, id=unavail_id, teacher=request.user)
    
    if request.method == 'POST':
        unavailability.delete()
        messages.success(request, "Unavailability removed!")
        return redirect('manage_unavailability')
    
    return render(request, 'scheduler/confirm_delete_unavailability.html', {'unavailability': unavailability})


def settings_view(request):
    user = request.user

    profile_form = ProfileForm(instance=user)
    password_form = PasswordChangeForm(user)

    if request.method == 'POST':
        if 'update_profile' in request.POST:
            profile_form = ProfileForm(request.POST, request.FILES, instance=user)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, 'Your profile has been updated!')
                return redirect('settings')

        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user) 
                messages.success(request, 'Your password was successfully updated!')
                return redirect('settings')
            else:
                messages.error(request, 'Please correct the error below.')

    context = {
        'profile_form': profile_form,
        'password_form': password_form
    }
    return render(request, 'scheduler/settings.html', context)


def course_list(request):
    courses = Course.objects.all()
    return render(request, 'scheduler/course_list.html', {'courses': courses})


def edit_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            return redirect('course_list')
    else:
        form = CourseForm(instance=course)
    
    return render(request, 'scheduler/add_course.html', {'form': form, 'is_edit': True})


def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        course.delete()
        return redirect('course_list')
    
    return render(request, 'scheduler/confirm_delete_course.html', {'course': course})


# ============= EXPORT HELPERS =============

def get_filtered_sessions_simple(request):
    """Simplified version to get sessions"""
    user = request.user
    selected_filiere_id = request.GET.get('filiere')
    
    if user.role == 'A':
        queryset = ScheduledSession.objects.all()
        
        if selected_filiere_id and selected_filiere_id != 'None':
            queryset = queryset.filter(course__filiere_id=selected_filiere_id)
            
    elif user.role == 'T':
        queryset = ScheduledSession.objects.filter(course__teacher=user)
        
        if selected_filiere_id and selected_filiere_id != 'None':
            queryset = queryset.filter(course__filiere_id=selected_filiere_id)
            
    else:
        if not user.student_group:
            return ScheduledSession.objects.none()
        
        queryset = ScheduledSession.objects.filter(
            Q(course__group=user.student_group) | 
            Q(course__filiere=user.student_group.filiere, course__group__isnull=True)
        )
    
    return queryset.select_related('course', 'room', 'course__teacher', 'course__filiere', 'course__group')


@login_required
def export_excel(request):
    """Export timetable as Excel"""
    sessions = get_filtered_sessions_simple(request)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Emploi du Temps"
    
    # Title
    ws['A1'] = "EMPLOI DU TEMPS UNIVERSITAIRE"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    
    # Info
    ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'] = f"Généré par: {request.user.get_full_name() or request.user.username}"
    ws['A4'] = f"Rôle: {request.user.get_role_display()}"
    
    # Add filter info
    selected_filiere_id = request.GET.get('filiere')
    if selected_filiere_id and selected_filiere_id != 'None':
        try:
            filiere = Filiere.objects.get(id=selected_filiere_id)
            ws['A5'] = f"Filière: {filiere.name} ({filiere.code})"
        except:
            pass
    
    # Headers
    ws['A7'] = "LISTE DES SÉANCES"
    ws['A7'].font = Font(bold=True, size=14)
    
    headers = ["Cours", "Type", "Enseignant", "Salle", "Jour", "Heure début", "Heure fin", "Groupe/Filière"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 9
    for session in sessions:
        if session.course.group:
            group_info = f"Groupe: {session.course.group.name}"
        else:
            group_info = f"Filière: {session.course.filiere.code}"
        
        ws.cell(row=row, column=1, value=session.course.name)
        ws.cell(row=row, column=2, value=session.course.get_session_type_display())
        ws.cell(row=row, column=3, value=session.course.teacher.get_full_name() or session.course.teacher.username)
        ws.cell(row=row, column=4, value=session.room.name)
        ws.cell(row=row, column=5, value=session.day)
        ws.cell(row=row, column=6, value=f"{session.start_hour}:00")
        ws.cell(row=row, column=7, value=f"{session.end_hour}:00")
        ws.cell(row=row, column=8, value=group_info)
        
        # Color by session type
        fill_color = "FFFFFF"
        
        if session.course.session_type == 'CM':
            fill_color = "E2EFDA"
        elif session.course.session_type == 'TD':
            fill_color = "FFF2CC"
        elif session.course.session_type == 'TP':
            fill_color = "DDEBF7"
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="emploi_du_temps.xlsx"'
    wb.save(response)
    return response